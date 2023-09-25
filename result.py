from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
max_row=ws.max_row
i=0
for row in range(2,max_row+1):
    hour=ws['C' + str(row)].value
    rate=ws['B' + str(row)].value
    if type(hour)!=str and type(rate)!=str:
        salary=hour*rate
        ws['D'+str(row)].value=salary
        print(salary)
    if salary>3000:
       i=i+1
       
    else:
       i=i+0
       
print('Amount of people with salary>3000 is =') 
print(i)
total=i

wb.save('result.xlsx')
wb.close()